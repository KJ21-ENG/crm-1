from datetime import datetime
from io import BytesIO
from unittest.mock import patch

import frappe
from openpyxl import load_workbook
from frappe.tests import IntegrationTestCase
from frappe.utils import add_days, getdate
from frappe.utils.file_manager import save_file
from frappe.utils.xlsxutils import make_xlsx

from crm.api.call_log import export_call_log_monthly_summary
from crm.api.dashboard import get_dashboard_data, get_today_celebrations
from crm.api.dashboard import mark_celebration_contacted
from crm.api.imports import (
	SUPPORTED_IMPORT_DOCTYPES,
	build_import_template_rows_for_test,
	download_import_template,
	get_bulk_import_status,
	start_bulk_import,
)


class IntegrationTestDashboardImportsAndExports(IntegrationTestCase):
	def setUp(self):
		super().setUp()
		frappe.set_user("Administrator")
		self.created_docs = []
		suffix = frappe.generate_hash(length=6)
		self.lead_status_name = f"Tmp Import Lead Status {suffix}"
		self.ticket_status_name = f"Tmp Import Ticket Status {suffix}"
		self.account_type_name = f"Tmp Import Account Type {suffix}"
		self.ticket_subject_name = f"Tmp Import Subject {suffix}"
		self._ensure_lead_status(self.lead_status_name)
		self._ensure_ticket_status(self.ticket_status_name)
		self._ensure_account_type(self.account_type_name)
		self._ensure_ticket_subject(self.ticket_subject_name)

	def tearDown(self):
		for doctype, name in reversed(self.created_docs):
			if frappe.db.exists(doctype, name):
				try:
					frappe.delete_doc(doctype, name, force=1)
				except Exception:
					pass
		frappe.set_user("Administrator")
		super().tearDown()

	def test_dashboard_payload_includes_today_celebrations(self):
		today = getdate()
		other_day = add_days(today, 1)

		birthday_customer = self._create_customer(
			{
				"customer_name": "Birthday Customer",
				"mobile_no": self._unique_mobile("91"),
				"date_of_birth": today,
			}
		)
		anniversary_customer = self._create_customer(
			{
				"customer_name": "Anniversary Customer",
				"mobile_no": self._unique_mobile("92"),
				"anniversary": today,
			}
		)
		self._create_customer(
			{
				"customer_name": "Other Day Customer",
				"mobile_no": self._unique_mobile("93"),
				"date_of_birth": other_day,
			}
		)

		payload = get_dashboard_data(view="daily")
		celebrations = payload.get("celebrations_today") or {}
		birthday_names = {row["name"] for row in celebrations.get("birthdays", [])}
		anniversary_names = {row["name"] for row in celebrations.get("anniversaries", [])}

		self.assertIn(birthday_customer.name, birthday_names)
		self.assertIn(anniversary_customer.name, anniversary_names)
		self.assertNotIn("Other Day Customer", {row.get("customer_name") for row in celebrations.get("birthdays", [])})

	def test_get_today_celebrations_returns_empty_without_customer_read_access(self):
		with patch("crm.api.dashboard.frappe.has_permission", return_value=False):
			payload = get_today_celebrations()
		self.assertEqual(payload["birthdays"], [])
		self.assertEqual(payload["anniversaries"], [])
		self.assertEqual(payload["pending_count"], 0)

	def test_mark_celebration_contacted_reduces_pending_count(self):
		today_customer = self._create_customer(
			{
				"customer_name": "Contacted Birthday Customer",
				"mobile_no": self._unique_mobile("94"),
				"date_of_birth": getdate(),
			}
		)

		initial = get_dashboard_data(view="daily").get("celebrations_today") or {}
		initial_pending = initial.get("pending_count", 0)
		updated = mark_celebration_contacted(today_customer.name, "Birthday")
		self.assertEqual(updated.get("pending_count"), max(initial_pending - 1, 0))
		birthday_row = next(row for row in updated.get("birthdays", []) if row["name"] == today_customer.name)
		self.assertTrue(birthday_row["contacted"])

		note_name = frappe.db.get_value(
			"FCRM Note",
			{
				"reference_doctype": "CRM Customer",
				"reference_docname": today_customer.name,
				"title": ["like", "Celebration Contacted::Birthday::%"],
			},
			"name",
		)
		if note_name:
			self.created_docs.append(("FCRM Note", note_name))

	def test_call_log_monthly_summary_can_be_filtered_to_single_month(self):
		self._create_call_log(
			{
				"id": f"CALL-{frappe.generate_hash(length=8)}",
				"from": "1111111111",
				"to": "9999999999",
				"type": "Incoming",
				"status": "Completed",
				"duration": 90,
				"employee": "Administrator",
				"start_time": datetime(2026, 1, 15, 10, 0, 0),
			}
		)
		self._create_call_log(
			{
				"id": f"CALL-{frappe.generate_hash(length=8)}",
				"from": "2222222222",
				"to": "9999999999",
				"type": "Outgoing",
				"status": "Completed",
				"duration": 120,
				"employee": "Administrator",
				"start_time": datetime(2026, 2, 18, 12, 0, 0),
			}
		)

		frappe.local.response = {}
		export_call_log_monthly_summary(file_format_type="CSV", month="2026-02")
		result = frappe.response.get("result", "")

		self.assertIn("Feb 2026", result)
		self.assertNotIn("Jan 2026", result)
		self.assertEqual(frappe.response.get("type"), "csv")

	def test_bulk_import_rejects_unsupported_doctype(self):
		with self.assertRaises(frappe.ValidationError):
			start_bulk_import("CRM Customer", "/private/files/not-supported.xlsx")

	def test_download_import_template_marks_required_headers_red(self):
		frappe.local.response = {}
		download_import_template("CRM Lead")
		workbook = load_workbook(BytesIO(frappe.response["filecontent"]))
		worksheet = workbook.active

		required_header_positions = {
			fieldname: index + 1
			for index, fieldname in enumerate(SUPPORTED_IMPORT_DOCTYPES["CRM Lead"])
			if frappe.get_meta("CRM Lead").get_field(fieldname)
			and frappe.get_meta("CRM Lead").get_field(fieldname).reqd
		}

		for column_index in required_header_positions.values():
			self.assertEqual(worksheet.cell(row=1, column=column_index).font.color.rgb, "00FF0000")

	def test_bulk_import_creates_insert_jobs_for_lead_and_ticket(self):
		lead_status = self._run_import(
			"CRM Lead",
			{
				"first_name": "Imported",
				"last_name": "Lead",
				"mobile_no": self._unique_mobile("81"),
				"lead_source": "On Call",
				"lead_category": "Direct",
				"account_type": self.account_type_name,
				"status": self.lead_status_name,
			},
		)
		lead_import_type = frappe.db.get_value("Data Import", lead_status["data_import_name"], "import_type")
		lead_name = frappe.db.get_value(
			"Data Import Log",
			{"data_import": lead_status["data_import_name"], "success": 1},
			"docname",
		)
		lead_customer = frappe.db.get_value("CRM Customer", {"mobile_no": lead_status["mobile_no"]}, "name")
		if lead_customer:
			self.created_docs.append(("CRM Customer", lead_customer))
		if lead_name:
			self.created_docs.append(("CRM Lead", lead_name))

		ticket_status = self._run_import(
			"CRM Ticket",
			{
				"ticket_subject": self.ticket_subject_name,
				"description": "Imported ticket description",
				"priority": "Medium",
				"issue_type": "General",
				"ticket_source": "On Call",
				"status": self.ticket_status_name,
			},
		)
		ticket_import_type = frappe.db.get_value("Data Import", ticket_status["data_import_name"], "import_type")
		ticket_name = frappe.db.get_value(
			"Data Import Log",
			{"data_import": ticket_status["data_import_name"], "success": 1},
			"docname",
		)
		if ticket_name:
			self.created_docs.append(("CRM Ticket", ticket_name))

		self.assertEqual(lead_import_type, "Insert New Records")
		self.assertEqual(ticket_import_type, "Insert New Records")
		self.assertTrue(lead_status["is_terminal"])
		self.assertTrue(ticket_status["is_terminal"])
		self.assertEqual(lead_status["status"], "Success")
		self.assertEqual(ticket_status["status"], "Success")
		self.assertTrue(lead_name)
		self.assertTrue(ticket_name)

	def _run_import(self, doctype, field_values):
		rows = build_import_template_rows_for_test(doctype)
		data_row = [""] * len(rows[0])
		for index, fieldname in enumerate(SUPPORTED_IMPORT_DOCTYPES[doctype]):
			if fieldname in field_values:
				data_row[index] = field_values[fieldname]
		rows[1] = data_row

		filename = f"{doctype.replace(' ', '_').lower()}_{frappe.generate_hash(length=6)}.xlsx"
		file_doc = save_file(filename, make_xlsx(rows, doctype).getvalue(), None, None, is_private=1)
		self.created_docs.append(("File", file_doc.name))

		status = start_bulk_import(doctype, file_doc.file_url)
		for _ in range(5):
			if status.get("is_terminal"):
				break
			status = get_bulk_import_status(status["data_import_name"])

		data_import_name = status["data_import_name"]
		self.created_docs.append(("Data Import", data_import_name))

		for imported_docname in frappe.get_all(
			"Data Import Log",
			filters={"data_import": data_import_name, "success": 1},
			pluck="docname",
		):
			if imported_docname:
				self.created_docs.append((doctype, imported_docname))

		if doctype == "CRM Lead":
			status["mobile_no"] = field_values.get("mobile_no")

		return status

	def _create_customer(self, values):
		doc = frappe.get_doc(
			{
				"doctype": "CRM Customer",
				"customer_name": values["customer_name"],
				"mobile_no": values["mobile_no"],
				"date_of_birth": values.get("date_of_birth"),
				"anniversary": values.get("anniversary"),
				"status": values.get("status") or "Active",
				"customer_source": "Direct",
			}
		).insert(ignore_permissions=True)
		self.created_docs.append(("CRM Customer", doc.name))
		return doc

	def _create_call_log(self, values):
		doc = frappe.get_doc({"doctype": "CRM Call Log", **values}).insert(ignore_permissions=True)
		self.created_docs.append(("CRM Call Log", doc.name))
		return doc

	def _ensure_lead_status(self, name):
		if frappe.db.exists("CRM Lead Status", name):
			return
		doc = frappe.get_doc(
			{
				"doctype": "CRM Lead Status",
				"lead_status": name,
				"color": "blue",
				"position": 999,
			}
		).insert(ignore_permissions=True)
		self.created_docs.append(("CRM Lead Status", doc.name))

	def _ensure_ticket_status(self, name):
		if frappe.db.exists("CRM Ticket Status", name):
			return
		doc = frappe.get_doc(
			{
				"doctype": "CRM Ticket Status",
				"ticket_status": name,
				"color": "orange",
				"position": 999,
			}
		).insert(ignore_permissions=True)
		self.created_docs.append(("CRM Ticket Status", doc.name))

	def _ensure_account_type(self, name):
		if frappe.db.exists("CRM Account Type", name):
			return
		doc = frappe.get_doc(
			{
				"doctype": "CRM Account Type",
				"account_type": name,
				"is_active": 1,
			}
		).insert(ignore_permissions=True)
		self.created_docs.append(("CRM Account Type", doc.name))

	def _ensure_ticket_subject(self, name):
		if frappe.db.exists("CRM Ticket Subject", name):
			return
		doc = frappe.get_doc(
			{
				"doctype": "CRM Ticket Subject",
				"subject_name": name,
				"is_active": 1,
			}
		).insert(ignore_permissions=True)
		self.created_docs.append(("CRM Ticket Subject", doc.name))

	def _unique_mobile(self, prefix):
		seed = datetime.utcnow().strftime("%H%M%S%f")[-8:]
		return f"{prefix}{seed}"[:10]
