import re
from io import BytesIO

import frappe
from frappe import _
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from frappe.core.doctype.data_import.data_import import (
	get_import_logs as core_get_import_logs,
	get_import_status as core_get_import_status,
)
from frappe.core.doctype.data_import.exporter import Exporter
from frappe.permissions import can_import
from frappe.utils.file_manager import save_file
from frappe.utils import cstr


SUPPORTED_IMPORT_DOCTYPES = {
	"CRM Lead": [
		"first_name",
		"last_name",
		"mobile_no",
		"alternative_mobile_no",
		"email",
		"organization",
		"job_title",
		"lead_source",
		"lead_category",
		"account_type",
		"status",
		"lead_owner",
		"referral_through",
		"client_id",
		"pod_id",
		"marital_status",
		"date_of_birth",
		"anniversary",
		"pan_card_number",
		"aadhaar_card_number",
	],
	"CRM Ticket": [
		"ticket_subject",
		"description",
		"priority",
		"issue_type",
		"ticket_source",
		"status",
		"assigned_to",
		"ticket_owner",
		"first_name",
		"last_name",
		"mobile_no",
		"email",
		"organization",
		"marital_status",
		"date_of_birth",
		"anniversary",
		"pan_card_number",
		"aadhaar_card_number",
	],
}

TERMINAL_IMPORT_STATUSES = {"Success", "Partial Success", "Error", "Timed Out"}
XLSX_FILENAME_PATTERN = re.compile(r"\.xlsx$", re.IGNORECASE)
TEMPLATE_DATA_ROW_START = 2
TEMPLATE_DATA_ROW_END = 250


def _validate_supported_doctype(doctype: str) -> str:
	if doctype not in SUPPORTED_IMPORT_DOCTYPES:
		frappe.throw(_("Bulk import is supported only for CRM Lead and CRM Ticket."))
	return doctype


def _validate_file_url(file_url: str) -> str:
	if not file_url:
		frappe.throw(_("An uploaded template file is required."))

	if not XLSX_FILENAME_PATTERN.search(file_url):
		frappe.throw(_("Only .xlsx template files are supported for bulk import."))

	file_doc = frappe.db.get_value(
		"File",
		{"file_url": file_url},
		["name", "file_url"],
		as_dict=True,
	)
	if not file_doc:
		frappe.throw(_("Uploaded template file could not be found."))

	return file_doc.file_url


def _build_template_rows(doctype: str) -> list[list[str]]:
	exporter = Exporter(
		doctype,
		export_fields={doctype: SUPPORTED_IMPORT_DOCTYPES[doctype]},
		export_data=False,
		file_type="Excel",
	)
	return exporter.get_csv_array_for_export()


def _get_required_template_fields(doctype: str) -> set[str]:
	meta = frappe.get_meta(doctype)
	required_fields = {
		fieldname
		for fieldname in SUPPORTED_IMPORT_DOCTYPES[doctype]
		if meta.get_field(fieldname) and meta.get_field(fieldname).reqd
	}

	if doctype == "CRM Ticket":
		required_fields.add("ticket_owner")

	return required_fields


def _get_allowed_user_entries(doctype: str):
	meta = frappe.get_meta(doctype)
	roles = {
		perm.role
		for perm in meta.permissions
		if getattr(perm, "write", 0) and perm.role not in {"All", "Guest"}
	}
	if not roles:
		return []

	user_names = frappe.get_all(
		"Has Role",
		filters={"role": ["in", list(roles)], "parenttype": "User"},
		pluck="parent",
	)
	user_names = sorted(set(user_names))
	users = frappe.get_all(
		"User",
		filters={"name": ["in", user_names], "enabled": 1},
		fields=["name", "full_name"],
		order_by="full_name asc, name asc",
	)
	user_map = {user["name"]: user for user in users}
	if frappe.session.user and frappe.session.user not in user_map:
		user_doc = frappe.db.get_value(
			"User",
			frappe.session.user,
			["name", "full_name"],
			as_dict=True,
		)
		if user_doc:
			users.insert(0, user_doc)

	entries = []
	for user in users:
		label = user["name"]
		if user.get("full_name") and user["full_name"] != user["name"]:
			label = f"{user['full_name']} ({user['name']})"
		entries.append({"label": label, "value": user["name"]})
	return entries


def _get_dropdown_config(doctype: str):
	meta = frappe.get_meta(doctype)
	user_entries = _get_allowed_user_entries(doctype)
	user_options = [entry["label"] for entry in user_entries]
	user_map = {entry["label"]: entry["value"] for entry in user_entries}
	for entry in user_entries:
		user_map[entry["value"]] = entry["value"]

	config = {}
	for fieldname in SUPPORTED_IMPORT_DOCTYPES[doctype]:
		field = meta.get_field(fieldname)
		if not field:
			continue

		options = None
		value_map = {}

		if fieldname in {"lead_owner", "assigned_to", "ticket_owner"}:
			options = user_options
			value_map = user_map
		elif field.fieldtype == "Select" and field.options:
			options = [option for option in cstr(field.options).split("\n") if option]
			value_map = {option: option for option in options}
		elif fieldname == "status":
			status_doctype = "CRM Lead Status" if doctype == "CRM Lead" else "CRM Ticket Status"
			options = frappe.get_all(status_doctype, pluck="name", order_by="position asc, name asc")
			value_map = {option: option for option in options}
		elif fieldname == "account_type":
			options = frappe.get_all("CRM Account Type", pluck="name", order_by="account_type asc")
			value_map = {option: option for option in options}
		elif fieldname == "ticket_subject":
			options = frappe.get_all("CRM Ticket Subject", pluck="name", order_by="subject_name asc")
			value_map = {option: option for option in options}

		if options:
			config[fieldname] = {
				"options": options,
				"value_map": value_map,
			}

	return config


def _get_date_template_fields(doctype: str):
	meta = frappe.get_meta(doctype)
	return {
		fieldname
		for fieldname in SUPPORTED_IMPORT_DOCTYPES[doctype]
		if meta.get_field(fieldname)
		and meta.get_field(fieldname).fieldtype in {"Date", "Datetime"}
	}


def _build_template_xlsx(doctype: str) -> bytes:
	rows = _build_template_rows(doctype)
	required_fields = _get_required_template_fields(doctype)
	dropdown_config = _get_dropdown_config(doctype)
	date_fields = _get_date_template_fields(doctype)
	default_ticket_owner = None
	if doctype == "CRM Ticket":
		user_entries = _get_allowed_user_entries(doctype)
		default_entry = next((entry for entry in user_entries if entry["value"] == frappe.session.user), None)
		default_ticket_owner = default_entry["label"] if default_entry else frappe.session.user

	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = doctype.replace(" ", "_")[:31]
	options_sheet = workbook.create_sheet("Options")
	options_sheet.sheet_state = "hidden"

	for row_index, row in enumerate(rows, start=1):
		for column_index, value in enumerate(row, start=1):
			cell = worksheet.cell(row=row_index, column=column_index, value=value)
			if row_index == 1:
				fieldname = SUPPORTED_IMPORT_DOCTYPES[doctype][column_index - 1]
				if fieldname in required_fields:
					cell.font = Font(name="Calibri", bold=True, color="FF0000")
				else:
					cell.font = Font(name="Calibri", bold=True)

	for column_index, fieldname in enumerate(SUPPORTED_IMPORT_DOCTYPES[doctype], start=1):
		config = dropdown_config.get(fieldname)
		if config:
			option_column = get_column_letter(column_index)
			for option_index, option in enumerate(config["options"], start=1):
				options_sheet.cell(row=option_index, column=column_index, value=option)

			formula = f"'Options'!${option_column}$1:${option_column}${len(config['options'])}"
			validation = DataValidation(type="list", formula1=formula, allow_blank=True)
			worksheet.add_data_validation(validation)
			validation.add(f"{get_column_letter(column_index)}{TEMPLATE_DATA_ROW_START}:{get_column_letter(column_index)}{TEMPLATE_DATA_ROW_END}")

		if fieldname in date_fields:
			column_letter = get_column_letter(column_index)
			date_validation = DataValidation(
				type="date",
				operator="between",
				formula1="DATE(1900,1,1)",
				formula2="DATE(2100,12,31)",
				allow_blank=True,
			)
			date_validation.promptTitle = "Date"
			date_validation.prompt = "Use the date picker or enter the date in YYYY-MM-DD format."
			date_validation.errorTitle = "Invalid date"
			date_validation.error = "Enter a valid date in YYYY-MM-DD format."
			worksheet.add_data_validation(date_validation)
			date_validation.add(f"{column_letter}{TEMPLATE_DATA_ROW_START}:{column_letter}{TEMPLATE_DATA_ROW_END}")
			for row_index in range(TEMPLATE_DATA_ROW_START, TEMPLATE_DATA_ROW_END + 1):
				worksheet.cell(row=row_index, column=column_index).number_format = "yyyy-mm-dd"

		if doctype == "CRM Ticket" and fieldname == "ticket_owner" and default_ticket_owner:
			for row_index in range(TEMPLATE_DATA_ROW_START, TEMPLATE_DATA_ROW_END + 1):
				worksheet.cell(row=row_index, column=column_index, value=default_ticket_owner)

	buffer = BytesIO()
	workbook.save(buffer)
	return buffer.getvalue()


def _format_import_messages(value):
	if not value:
		return []

	if isinstance(value, list):
		return [str(item) for item in value if item]

	if isinstance(value, str):
		try:
			parsed = frappe.parse_json(value)
		except Exception:
			return [value]

		if isinstance(parsed, list):
			out = []
			for item in parsed:
				if isinstance(item, dict):
					out.append(item.get("message") or frappe.as_json(item))
				elif item:
					out.append(str(item))
			return out

		if isinstance(parsed, dict):
			return [parsed.get("message") or frappe.as_json(parsed)]

	return [str(value)]


def _build_processed_import_file(doctype: str, file_url: str) -> str:
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	workbook = load_workbook(filename=file_doc.get_full_path(), data_only=True)
	worksheet = workbook.active
	headers = _build_template_rows(doctype)[0]
	dropdown_config = _get_dropdown_config(doctype)
	user_value_map = {}
	for fieldname in {"lead_owner", "assigned_to", "ticket_owner"}:
		user_value_map.update(dropdown_config.get(fieldname, {}).get("value_map", {}))

	processed_rows = [headers]
	for row_index in range(TEMPLATE_DATA_ROW_START, worksheet.max_row + 1):
		row_values = []
		has_meaningful_data = False

		for column_index, fieldname in enumerate(SUPPORTED_IMPORT_DOCTYPES[doctype], start=1):
			value = worksheet.cell(row=row_index, column=column_index).value
			if isinstance(value, str):
				value = value.strip()

			if fieldname in {"lead_owner", "assigned_to", "ticket_owner"} and value:
				value = user_value_map.get(value, value)

			if fieldname == "ticket_owner" and not value:
				value = frappe.session.user

			if fieldname != "ticket_owner" and value not in (None, ""):
				has_meaningful_data = True

			row_values.append(value)

		if has_meaningful_data:
			processed_rows.append(row_values)

	if len(processed_rows) == 1:
		frappe.throw(_("No filled rows were found in the uploaded template."))

	filename = f"{doctype.replace(' ', '_')}_Processed_{frappe.generate_hash(length=6)}.xlsx"
	processed_file = save_file(filename, _build_processed_xlsx(processed_rows), None, None, is_private=1)
	return processed_file.file_url


def _build_processed_xlsx(rows: list[list[str]]) -> bytes:
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = "Data"

	for row in rows:
		worksheet.append(row)

	buffer = BytesIO()
	workbook.save(buffer)
	return buffer.getvalue()


def _can_access_import(doc) -> bool:
	return doc.owner == frappe.session.user or "System Manager" in frappe.get_roles()


@frappe.whitelist()
def download_import_template(doctype: str):
	doctype = _validate_supported_doctype(doctype)
	can_import(doctype, raise_exception=True)

	filename = f"{doctype.replace(' ', '_')}_Import_Template"
	content = _build_template_xlsx(doctype)
	frappe.response["filename"] = f"{filename}.xlsx"
	frappe.response["filecontent"] = content
	frappe.response["type"] = "binary"
	frappe.response["doctype"] = doctype


@frappe.whitelist()
def start_bulk_import(doctype: str, file_url: str):
	doctype = _validate_supported_doctype(doctype)
	can_import(doctype, raise_exception=True)
	file_url = _validate_file_url(file_url)
	processed_file_url = _build_processed_import_file(doctype, file_url)

	data_import = frappe.get_doc(
		{
			"doctype": "Data Import",
			"reference_doctype": doctype,
			"import_type": "Insert New Records",
			"import_file": processed_file_url,
			"mute_emails": 1,
			"submit_after_import": 0,
		}
	)
	data_import.insert(ignore_permissions=True)
	frappe.db.commit()
	data_import.start_import()

	status = get_bulk_import_status(data_import.name)
	status["started"] = True
	return status


@frappe.whitelist()
def get_bulk_import_status(data_import_name: str):
	if not data_import_name:
		frappe.throw(_("Data Import name is required."))

	doc = frappe.get_doc("Data Import", data_import_name)
	if not _can_access_import(doc):
		frappe.throw(_("Not permitted to access this import."), frappe.PermissionError)

	_validate_supported_doctype(doc.reference_doctype)

	status = core_get_import_status(data_import_name)
	logs = core_get_import_logs(data_import_name)

	failed_logs = []
	for log in logs:
		if log.get("success"):
			continue
		failed_logs.append(
			{
				"row_indexes": frappe.parse_json(log.get("row_indexes") or "[]"),
				"messages": _format_import_messages(log.get("messages")),
				"exception": log.get("exception"),
			}
		)

	return {
		"data_import_name": data_import_name,
		"doctype": doc.reference_doctype,
		"status": status.get("status"),
		"success_count": status.get("success", 0),
		"failed_count": status.get("failed", 0),
		"total_records": status.get("total_records", 0),
		"is_terminal": status.get("status") in TERMINAL_IMPORT_STATUSES,
		"failed_logs": failed_logs,
	}


def build_import_template_rows_for_test(doctype: str):
	"""Test helper for building template rows without relying on HTTP response state."""
	return _build_template_rows(_validate_supported_doctype(doctype))
